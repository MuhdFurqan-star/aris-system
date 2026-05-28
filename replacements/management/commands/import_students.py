"""
Django management command to import students and lecturers from the
PELAJAR LATIHAN INDUSTRI Excel file.

Usage:
    python manage.py import_students <path_to_excel>

What it does:
  1. Reads all 35 students from the DCS sheet.
  2. Deduplicates lecturer names from NAMA MENTOR + NAMA PENYELIA LI columns.
  3. Creates lecturer accounts (LC0001, LC0002 ...) -- skips existing ones.
  4. Creates student accounts (username = student ID) -- skips existing ones.
  5. Enrolls every student into ALL Semester-6 classes (6A, 6B).
  6. Spreads all created lecturers evenly across every unassigned subject system-wide.
"""

import re
from openpyxl import load_workbook
from django.core.management.base import BaseCommand
from django.contrib.auth.models import User
from replacements.models import UserProfile, Semester, Subject, StudentEnrollment


# -- Honorific titles to strip from lecturer names --------------------------
TITLES = {
    'pn', 'puan', 'dr', 'dr.', 'en', 'en.', 'encik',
    'prof', 'prof.', 'tuan', 'mdm', 'mdm.',
    'hj', 'hj.', 'hjh', 'hjh.', 'dato', "dato'",
}


def strip_title(name: str) -> str:
    """Remove leading honorific titles from a name string."""
    words = name.strip().split()
    while words and words[0].lower().rstrip('.') in TITLES:
        words = words[1:]
    return ' '.join(words)


def normalise_name(name: str) -> str:
    """
    Canonical uppercase version of a lecturer name used as dedup key.
    * Strips titles
    * Replaces 'BT'/'Bt' abbreviation -> 'BINTI'
    * Removes trailing dots from words (MOHD. -> MOHD)
    * Collapses whitespace
    """
    # Insert space after dot+letter (e.g. "DR.ZALINA" -> "DR. ZALINA")
    clean = re.sub(r'\.([A-Za-z])', r'. \1', name.strip())
    clean = strip_title(clean)
    # Replace standalone Bt/BT with BINTI
    clean = re.sub(r'\bBT\b', 'BINTI', clean, flags=re.IGNORECASE)
    # Remove dots inside words (MOHD. -> MOHD)
    clean = re.sub(r'\.', '', clean)
    # Collapse spaces
    clean = re.sub(r'\s+', ' ', clean).strip().upper()
    return clean


def dedup_key(name: str) -> str:
    """
    Lightweight dedup key: first-word + last-word (uppercase, no dots).
    Handles spelling variants like 'Mansor' vs 'Mansur' where the canonical
    first and last words still match.
    """
    norm = normalise_name(name)
    words = norm.split()
    if len(words) >= 2:
        return f"{words[0]}_{words[-1]}"
    return norm


def make_email(full_name: str) -> str:
    """firstname.lastname@kpm.edu.my  (first word + last word, lowercase alpha only)."""
    words = full_name.strip().split()
    first = re.sub(r'[^a-z]', '', words[0].lower()) if words else 'user'
    last  = re.sub(r'[^a-z]', '', words[-1].lower()) if len(words) > 1 else 'user'
    return f"{first}.{last}@kpm.edu.my"


def unique_email(base: str, seen: set) -> str:
    """Append a counter suffix if the email is already taken."""
    email = base
    counter = 1
    local, domain = base.split('@')
    while email in seen or User.objects.filter(email=email).exists():
        email = f"{local}{counter}@{domain}"
        counter += 1
    seen.add(email)
    return email


def unique_username(base: str, seen: set) -> str:
    """Append a counter suffix if the username is already taken."""
    username = base
    counter = 1
    while username in seen or User.objects.filter(username=username).exists():
        username = f"{base}{counter}"
        counter += 1
    seen.add(username)
    return username


# -- Command -----------------------------------------------------------------

class Command(BaseCommand):
    help = 'Import students and lecturers from Latihan Industri Excel file'

    def add_arguments(self, parser):
        parser.add_argument(
            'excel_path',
            type=str,
            help='Full path to the .xlsx file',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Preview what would be created without saving anything',
        )
        parser.add_argument(
            '--course',
            type=str,
            default='CS',
            choices=['CS', 'LH', 'ACC', 'BS'],
            help='Course code to tag imported users with (default: CS)',
        )

    def handle(self, *args, **options):
        excel_path = options['excel_path']
        dry_run    = options['dry_run']
        course     = options['course']

        if dry_run:
            self.stdout.write(self.style.WARNING('DRY RUN -- nothing will be saved\n'))

        self.stdout.write(f'Course tag: {course}\n')

        # -- 1. Parse Excel ---------------------------------------------------
        self.stdout.write('Reading Excel file …')
        wb = load_workbook(excel_path, data_only=True)
        ws = wb['DCS']

        students_data   = []   # list of dicts
        raw_mentor_map  = {}   # dedup_key -> raw canonical name
        raw_super_map   = {}   # dedup_key -> raw canonical name

        for row in ws.iter_rows(min_row=8, values_only=True):
            bil, name, sid, nkp, mentor, supervisor = (
                row[1], row[2], row[3], row[4], row[5], row[6]
            )
            if not name or not sid:
                continue

            s_name = str(name).strip()
            s_id   = str(sid).strip()
            m_raw  = str(mentor).strip()   if mentor   else ''
            sv_raw = str(supervisor).strip() if supervisor else ''

            students_data.append({
                'name':       s_name,
                'student_id': s_id,
                'mentor_raw': m_raw,
                'super_raw':  sv_raw,
            })

            # Prefer ALL-CAPS version (supervisor column) as canonical name
            if m_raw:
                k = dedup_key(m_raw)
                # Keep whichever version is already uppercase (more official)
                existing = raw_mentor_map.get(k, '')
                raw_mentor_map[k] = m_raw if (m_raw == m_raw.upper() and existing != m_raw.upper()) else existing or m_raw

            if sv_raw:
                k = dedup_key(sv_raw)
                existing = raw_super_map.get(k, '')
                raw_super_map[k] = sv_raw if (sv_raw == sv_raw.upper()) else existing or sv_raw

        self.stdout.write(f'  Found {len(students_data)} students')

        # Merge mentor + supervisor unique lecturers
        all_lec_map = {}  # dedup_key -> canonical raw name
        for k, v in raw_super_map.items():
            all_lec_map[k] = v          # supervisor takes precedence (all-caps)
        for k, v in raw_mentor_map.items():
            if k not in all_lec_map:
                all_lec_map[k] = v

        self.stdout.write(f'  Found {len(all_lec_map)} unique lecturers after deduplication\n')

        # -- 2. Get Semester 6 classes ----------------------------------------
        sem6_classes = list(Semester.objects.filter(name='6'))
        if not sem6_classes:
            # Fallback: try name contains '6'
            sem6_classes = list(Semester.objects.filter(name__icontains='6'))
        if not sem6_classes:
            self.stderr.write(self.style.ERROR(
                'Could not find Semester 6! Please check semester names in the database.'
            ))
            return

        self.stdout.write(f'Target semesters for enrollment:')
        for s in sem6_classes:
            self.stdout.write(f'  * Semester {s.name} - Class {s.class_name}')
        self.stdout.write('')

        # -- 3. Determine next LC ID number -----------------------------------
        existing_lc = UserProfile.objects.filter(
            user_type='lecturer',
            employee_id__regex=r'^LC\d+$',
        ).values_list('employee_id', flat=True)

        max_lc = 0
        for lc in existing_lc:
            try:
                num = int(lc[2:])
                max_lc = max(max_lc, num)
            except ValueError:
                pass

        # -- 4. Create lecturers ----------------------------------------------
        self.stdout.write('-' * 55)
        self.stdout.write('CREATING LECTURERS')
        self.stdout.write('-' * 55)

        used_emails    = set()
        used_usernames = set()
        created_lecs   = 0
        skipped_lecs   = 0
        lec_id_map     = {}   # dedup_key -> User object (for reporting)

        for key, raw_name in sorted(all_lec_map.items()):
            clean = normalise_name(raw_name)   # UPPERCASE normalised
            words = clean.split()
            first_name = words[0].title() if words else clean
            last_name  = ' '.join(words[1:]).title() if len(words) > 1 else ''

            # Check if a lecturer with same clean name already exists
            existing_user = User.objects.filter(
                userprofile__user_type='lecturer',
                first_name__iexact=first_name,
                last_name__iexact=last_name,
            ).first()

            if existing_user:
                lp = existing_user.userprofile
                self.stdout.write(
                    f'  [SKIP]   {clean} -> already exists as {lp.employee_id}'
                )
                lec_id_map[key] = existing_user
                skipped_lecs += 1
                continue

            max_lc += 1
            lc_id  = f'LC{max_lc:04d}'
            email  = unique_email(make_email(clean), used_emails)
            uname  = unique_username(lc_id, used_usernames)

            if not dry_run:
                user = User.objects.create_user(
                    username   = uname,
                    email      = email,
                    first_name = first_name,
                    last_name  = last_name,
                    password   = lc_id,
                )
                UserProfile.objects.create(
                    user        = user,
                    user_type   = 'lecturer',
                    employee_id = lc_id,
                    course      = course,
                )
                lec_id_map[key] = user

            self.stdout.write(
                f'  [CREATE] {clean:<45} -> {lc_id}  |  {email}'
            )
            created_lecs += 1

        self.stdout.write(
            f'\n  Lecturers: {created_lecs} created, {skipped_lecs} skipped\n'
        )

        # -- 5. Create students -----------------------------------------------
        self.stdout.write('-' * 55)
        self.stdout.write('CREATING STUDENTS')
        self.stdout.write('-' * 55)

        created_stds = 0
        skipped_stds = 0

        for s in students_data:
            sid       = s['student_id']
            full_name = s['name']

            if UserProfile.objects.filter(student_id=sid).exists():
                self.stdout.write(f'  [SKIP]   {full_name} ({sid})')
                skipped_stds += 1
                continue

            words      = full_name.split()
            first_name = words[0].title() if words else full_name
            last_name  = ' '.join(words[1:]).title() if len(words) > 1 else ''
            email      = unique_email(make_email(full_name), used_emails)
            uname      = unique_username(sid, used_usernames)

            if not dry_run:
                user = User.objects.create_user(
                    username   = uname,
                    email      = email,
                    first_name = first_name,
                    last_name  = last_name,
                    password   = sid,
                )
                UserProfile.objects.create(
                    user       = user,
                    user_type  = 'student',
                    student_id = sid,
                    course     = course,
                )
                # Enroll in every Semester-6 class
                for sem in sem6_classes:
                    StudentEnrollment.objects.get_or_create(
                        student  = user,
                        semester = sem,
                    )

            self.stdout.write(
                f'  [CREATE] {full_name:<45} ({sid})  |  {email}'
            )
            created_stds += 1

        self.stdout.write(
            f'\n  Students: {created_stds} created, {skipped_stds} skipped\n'
        )

        # -- 6. Assign lecturers to all unassigned subjects -------------------
        if dry_run:
            self.stdout.write('-' * 55)
            self.stdout.write('SUBJECT ASSIGNMENT (skipped in dry-run)')
            self.stdout.write('-' * 55)
        else:
            unassigned = list(Subject.objects.filter(lecturer__isnull=True))
            all_lec_users = list(
                User.objects.filter(userprofile__user_type='lecturer')
                            .order_by('userprofile__employee_id')
            )

            self.stdout.write('-' * 55)
            self.stdout.write(
                f'ASSIGNING LECTURERS TO SUBJECTS  '
                f'({len(unassigned)} unassigned, {len(all_lec_users)} lecturers)'
            )
            self.stdout.write('-' * 55)

            if all_lec_users and unassigned:
                for i, subject in enumerate(unassigned):
                    lecturer = all_lec_users[i % len(all_lec_users)]
                    subject.lecturer = lecturer
                    subject.save()
                    self.stdout.write(
                        f'  [ASSIGN] {subject.subject_code:<20} '
                        f'{subject.subject_name[:35]:<35} -> '
                        f'{lecturer.get_full_name()}'
                    )
            else:
                self.stdout.write('  Nothing to assign.')

        # -- Summary ----------------------------------------------------------
        self.stdout.write('\n' + '=' * 55)
        if dry_run:
            self.stdout.write(self.style.WARNING(
                f'DRY RUN complete. No data was saved.\n'
                f'  Would create: {created_lecs} lecturers, {created_stds} students'
            ))
        else:
            self.stdout.write(self.style.SUCCESS(
                f'Import complete!\n'
                f'  Lecturers : {created_lecs} created, {skipped_lecs} skipped\n'
                f'  Students  : {created_stds} created, {skipped_stds} skipped\n'
                f'  Enrolled  : {created_stds} students into '
                f'{len(sem6_classes)} Semester-6 class(es)\n'
                f'  Assigned  : subjects distributed among all lecturers'
            ))
        self.stdout.write('=' * 55)
